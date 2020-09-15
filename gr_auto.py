
# Writing to an excel  
# sheet using Python 
import xlwt 
from xlwt import Workbook 
from openpyxl import *
  
# Workbook is created 
wb = Workbook() 
wb=load_workbook("GRimport.xlsx")
ws=wb["Grocery Rescue"]



#Get date
date = input("Enter input date in m/d/yyyy format:") 
donor_inc = 6
agency_inc = 67
  
# Writes date to document













def get_drop_off_info():
	global donor_inc
	global agency_inc
	while True:
		pro_tot = 0
		meat_tot = 0
		dairy_tot = 0
		mix_tot = 0
		nonfood_tot = 0

		donor = input("What is the donor ID? If no donors remain type none: ")
		while(donor != "none"):
			donorcell = ws.cell(donor_inc, 1)
			donorcell.value = donor
			pickupcell = ws.cell(donor_inc, 4)
			pickupcell.value = "Pickup"


			produce = int(input("Produce: ") or 0)
			pro_tot += produce

			meat = int(input("Meat: ") or 0)
			meat_tot += meat

			dairy = int(input("Dairy: ") or 0)
			dairy_tot += dairy

			mix = int(input("Mix: ") or 0)
			mix_tot += mix

			nonfood = int(input("Nonfood: ") or 0)
			nonfood_tot += nonfood

			producecell = ws.cell(donor_inc, 6)
			meatcell = ws.cell(donor_inc, 7)
			dairycell = ws.cell(donor_inc, 8)
			mixcell = ws.cell(donor_inc, 9)
			nonfoodcell = ws.cell(donor_inc, 10)

			producecell.value = produce
			meatcell.value = meat
			dairycell.value = dairy
			mixcell.value = mix
			nonfoodcell.value = nonfood
			donor = input("What is the donor ID? If no donors remain type none: ")
			donor_inc += 1

		agency = input("What is the agency ID?")
		agencycell = ws.cell(agency_inc, 1)
		agencycell.value = agency

		delivercell = ws.cell(agency_inc, 4)
		delivercell.value = "Deliver"

		producecell = ws.cell(agency_inc, 6)
		meatcell = ws.cell(agency_inc, 7)
		dairycell = ws.cell(agency_inc, 8)
		mixcell = ws.cell(agency_inc, 9)
		nonfoodcell = ws.cell(agency_inc, 10)

		producecell.value = pro_tot
		meatcell.value = meat_tot
		dairycell.value = dairy_tot
		mixcell.value = mix_tot
		nonfoodcell.value = nonfood_tot
		agency_inc += 1
		more_sheets = input("More sheets? (y or n)")
		if(more_sheets == "n"):
			break

get_drop_off_info()
wb.save("pyexcel.xlsx")
  

